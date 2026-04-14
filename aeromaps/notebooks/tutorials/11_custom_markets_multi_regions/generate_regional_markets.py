"""
Read custom_data_excel/Classeur1.xlsx and generate one markets.yaml per region.

Output layout:
    data/<REGION>/markets.yaml   — one entry per aircraft subcategory (market)
    data/load_factor.yaml        — global load-factor values (not per market)

Load-factor values should be set on the process directly, e.g.:
    process.parameters.load_factor_2025 = ...
    process.parameters.load_factor_2045 = ...
    process.parameters.load_factor_2050 = ...
"""

import pathlib

import openpyxl
import yaml

EXCEL_PATH = pathlib.Path(__file__).parent / "custom_data_excel" / "Classeur1.xlsx"
DATA_DIR = pathlib.Path(__file__).parent / "data"
LOAD_FACTOR_SHEET = "load_factor"

# Full names for each aircraft-subcategory market id
MARKET_NAMES = {
    "NBI": "Narrow Body Intermediate",
    "NBL": "Narrow Body Large",
    "NBS": "Narrow Body Small",
    "RJ": "Regional Jet",
    "RT": "Regional Turboprop",
    "WBL": "Wide Body Large",
    "WBSI": "Wide Body Small International",
}

# All aircraft-subcategory markets are passenger / RPK
TRAFFIC_TYPE = "passenger"
TRAFFIC_UNIT = "RPK"


def read_matrix(ws):
    """Return {region: {market: value}} from a region×market sheet."""
    rows = list(ws.iter_rows(values_only=True))
    markets = [c for c in rows[0][1:] if c is not None]
    result = {}
    for row in rows[1:]:
        region = row[0]
        if region is None:
            continue
        result[region] = {m: row[i + 1] for i, m in enumerate(markets)}
    return result


def read_load_factor(ws):
    """Return {year: value} from the load_factor sheet."""
    rows = list(ws.iter_rows(values_only=True))
    result = {}
    for row in rows[1:]:
        year, value = row[0], row[1]
        if year is not None:
            result[int(year)] = value
    return result


def build_market_entry(
    market_id, ask_2019, ask_2025, cagr_ask, fleet_productivity_2025, fleet_productivity_2045
):
    return {
        "name": MARKET_NAMES.get(market_id, market_id),
        "traffic_type": TRAFFIC_TYPE,
        "traffic_unit": TRAFFIC_UNIT,
        "inputs": {
            "initial": {
                "ask_2019": ask_2019,
                "ask_2025": ask_2025,
                "fleet_productivity_2025": fleet_productivity_2025,
                "fleet_productivity_2045": fleet_productivity_2045,
            },
            "growth": {
                "cagr_reference_periods": [],
                "cagr_reference_periods_values": [cagr_ask],
            },
        },
    }


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    ask_2019_m = read_matrix(wb["ask_2019"])
    ask_2025_m = read_matrix(wb["ask_2025"])
    cagr_ask_m = read_matrix(wb["cagr_ask"])
    prod_2025_m = read_matrix(wb["fleet_profuctivity_2025"])  # workbook typo kept
    prod_2045_m = read_matrix(wb["fleet_productivity_2045"])

    regions = list(ask_2019_m.keys())
    markets = list(ask_2019_m[regions[0]].keys())

    for region in regions:
        region_dir = DATA_DIR / region
        region_dir.mkdir(parents=True, exist_ok=True)

        yaml_data = {}
        for market_id in markets:
            yaml_data[market_id] = build_market_entry(
                market_id,
                ask_2019=ask_2019_m[region][market_id],
                ask_2025=ask_2025_m[region][market_id],
                cagr_ask=cagr_ask_m[region][market_id],
                fleet_productivity_2025=prod_2025_m[region][market_id],
                fleet_productivity_2045=prod_2045_m[region][market_id],
            )

        out_path = region_dir / "markets.yaml"
        with open(out_path, "w") as f:
            yaml.dump(yaml_data, f, default_flow_style=False, sort_keys=False)
        print(f"Written: {out_path}")

    # --- load factor (global, not per market) ---
    lf = read_load_factor(wb[LOAD_FACTOR_SHEET])
    lf_path = DATA_DIR / "load_factor.yaml"
    with open(lf_path, "w") as f:
        yaml.dump(lf, f, default_flow_style=False)
    print(f"\nLoad factor (global) → {lf_path}")
    print("Set on process in notebook:")
    for year, value in lf.items():
        print(f"  process.parameters.load_factor_{year} = {value}")


if __name__ == "__main__":
    main()
